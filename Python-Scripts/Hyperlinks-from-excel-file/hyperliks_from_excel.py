import openpyxl
import sys
from tqdm import tqdm

def space(text):
    return text.replace('%20', ' ').replace('/', '\\')

wb = openpyxl.load_workbook("\\\\cpts-007\\backup\\IvanL\\Tunisia\\Vessels\\Inspections\\Hannibal PV's (Update 2019).xlsx")
ws = wb['Thorough ']

f = open('hyperlinks123.txt', 'a')

for i in range(0, 10000):


    for j in range(0, 10000):

        if j % 2 == 0:
            try:
                if ws.cell(row=i, column=j).value == 'yes' or ws.cell(row=i, column=j).value == 'Yes' or ws.cell(row=i, column=j).value == 'YES':
                    hlink = ws.cell(row=i, column=j).hyperlink.target
                    f.write('\\\\cpts-007\\backup\\IvanL\\Tunisia\\Vessels\\Inspections\\{}'.format(space(hlink)))
                    f.write("\n")

            except Exception as e:
                print("On ({},{}) error occupied: '{}'!".format(i, j, e))

f.close()

# %20 is space in URL decoding
